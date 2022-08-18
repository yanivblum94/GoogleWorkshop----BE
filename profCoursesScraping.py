import sys
from subprocess import check_call
import xlsxwriter
import os
from openpyxl import load_workbook
from openpyxl import Workbook

class Course:
    def __init__(self, name, number):
        self.name = name
        self.number = number
    

class Lecturer:
    def __init__(self, name, teach_type):
        self.name = name
        self.teach_type = teach_type
        self.email = ""
        self.website = ""
    
def install(package):
    check_call([sys.executable, "-m", "pip", "install", package])

def InitSelenium():
    try:
        import selenium
        import webdriver_manager
    except ImportError:
        install("selenium")
        install("webdriver-manager")
    finally:
        import selenium
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.firefox.options import Options
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support.ui import Select
        from selenium.webdriver.common.by import By
        import webdriver_manager
        from webdriver_manager.chrome import ChromeDriverManager
        from webdriver_manager.firefox import GeckoDriverManager
        global Select, By, Keys
        try:
            path = ChromeDriverManager().install()
            chrome_options = webdriver.ChromeOptions()
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--profile-directory=Default')
            driver = webdriver.Chrome(executable_path=path, options=chrome_options)
            return driver
        except:
            path = GeckoDriverManager().install()
            firefox_options = webdriver.FirefoxOptions()
            firefox_options.add_argument('--headless')
            firefox_options.add_argument('--profile-directory=Default')
            firefox_options.add_argument('--no-sandbox')
            firefox_options.add_argument('--disable-dev-shm-usage')
            driver = webdriver.Firefox(executable_path=path, options=firefox_options)
            return driver


def findAllCourses(driver, dep):
    faculty_field = driver.find_element_by_id(dep)
    select_option = Select(faculty_field)
    select_option.select_by_index(1)
    search_button = driver.find_element_by_id("search1")
    search_button.click()
    
def scrape_courses(driver, dep):
    page = 0
    row = 0
    more_pages = True
    wb = Workbook()
    worksheet = wb.active
    wb.save("Results.xlsx")

    while (more_pages):
        courses = {}
        driver.get("http://www.ims.tau.ac.il/tal/kr/search_p.aspx")
        faculty_field = driver.find_element_by_id(dep)
        select_option = Select(faculty_field)
        select_option.select_by_index(1)
        search_button = driver.find_element_by_id("search1")
        search_button.click()
        for i in range(page):
            try:
                next_button = driver.find_element_by_id("next")
                next_button.click()
            except:
                more_pages = False
                break
        table_rows = driver.find_elements_by_xpath("//*[@id='frmgrid']/table[2]/tbody/tr")
        i = 3
        course = None
        while i < len(table_rows):
            try:
                course_num = driver.find_elements_by_xpath("//*[@id='frmgrid']/table[2]/tbody/tr[" + str(i) + "]/td[1]")[0].text[0:9]
                course_name = driver.find_elements_by_xpath("//*[@id='frmgrid']/table[2]/tbody/tr[" + str(i) + "]/td[2]")[0].text
            except:
                continue
            if course == None or course_name != course.name:
                course = Course(course_name, course_num)
            course_type = ""
            i += 3
            while True:
                try:
                    if driver.find_element_by_xpath("//*[@id='frmgrid']/table[2]/tbody/tr[" + str(i+2) + "]").get_attribute("class") == "listtds kotcol":
                        break
                except:
                    break
                exists = False
                lecturer_name = driver.find_elements_by_xpath("//*[@id='frmgrid']/table[2]/tbody/tr[" + str(i) + "]/td[1]")[0].text
                if (lecturer_name == "") or (lecturer_name.isspace()):
                    i += 1
                    continue
                if course_type == "":
                    course_type = driver.find_elements_by_xpath("//*[@id='frmgrid']/table[2]/tbody/tr[" + str(i) + "]/td[2]")[0].text
                lecturer = Lecturer(lecturer_name, course_type)
                old_lecturers = courses.get(course)
                if old_lecturers == None:
                    courses[course] = []
                else:
                    for old_lecturer in old_lecturers:
                        if old_lecturer.name == lecturer.name:
                            exists = True
                if exists:
                    i += 1
                    continue
                courses[course].append(lecturer)
                i += 1
            i += 3
        print("starting writing" + str(row))
        row += get_email_and_website(wb, driver, courses, row)
        page += 1


def get_email_and_website(wb, driver, courses, row):
    worksheet = wb.active
    new_row = row
    if row == 0:
        data = ['course name', 'course number', 'lecturer name', 'type of lesson', 'email', 'website']
        worksheet.append(data)
        new_row += 1
    driver.get("https://www.tau.ac.il/tau/index")
    for course in courses:
        for lecturer in courses.get(course):
            try:
                first_name_field = driver.find_element_by_id("edit-first-name")
                last_name_field = driver.find_element_by_id("edit-last-name")
                first_name = lecturer.name.split(" ")[-1]
                last_name = lecturer.name.split(" ")[1:-1]
                last_name = " ".join(last_name)
                first_name_field.send_keys(first_name)
                last_name_field.send_keys(last_name)
                last_name_field.send_keys(Keys.ENTER)
                social = driver.find_elements_by_class_name("tau-person-social-network")
                try:
                    email = driver.find_elements_by_xpath("//*[@id='tau-person-results-table']/tbody/tr/td[5]/div[1]/a")[0].get_attribute('href')[7:]
                    lecturer.email = email
                except:
                    driver.find_element_by_id("edit-first-name").clear()
                    driver.find_element_by_id("edit-last-name").clear()
                    data = [str(course.name), str(course.number), str(lecturer.name), str(lecturer.teach_type), str(lecturer.email), str(lecturer.website)]
                    worksheet.append(data)
                    new_row += 1
                    continue
                if len(social) > 1:
                    try:
                        website = driver.find_elements_by_xpath("//*[@id='tau-person-results-table']/tbody/tr/td[5]/div[2]/a")[0].get_attribute('href')
                        lecturer.website = website
                    except:
                        driver.find_element_by_id("edit-first-name").clear()
                        driver.find_element_by_id("edit-last-name").clear()
                        data = [str(course.name), str(course.number), str(lecturer.name), str(lecturer.teach_type), str(lecturer.email), str(lecturer.website)]
                        worksheet.append(data)
                        new_row += 1
                        continue
                data = [str(course.name), str(course.number), str(lecturer.name), str(lecturer.teach_type), str(lecturer.email), str(lecturer.website)]
                worksheet.append(data)
                new_row += 1
                driver.find_element_by_id("edit-first-name").clear()
                driver.find_element_by_id("edit-last-name").clear()
                wb.save("Results.xlsx")
            except:
                wb.save("Results.xlsx")
    return new_row

def alfon_adding():
    wb = load_workbook("Results_try_try.xlsx")
    driver = InitSelenium()
    driver.get("https://www.tau.ac.il/tau/index")
    ws = wb.active
    i = 0
    for row in ws.rows:
        if i == 0:
            i += 1
            continue
        else:
            if row[4].value == None:
                name = row[2].value
                name = name.split(" ")
                last_name = name[1]
                last_name_field = driver.find_element_by_id("edit-last-name") 
                last_name_field.send_keys(last_name)
                last_name_field.send_keys(Keys.ENTER)
                if len(driver.find_elements_by_class_name("tau-person-social-td")) != 1:
                    driver.find_element_by_id("edit-last-name").clear()
                    continue
                social = driver.find_elements_by_class_name("tau-person-social-network")
                try:
                    email = driver.find_elements_by_xpath("//*[@id='tau-person-results-table']/tbody/tr/td[5]/div[1]/a")[0].get_attribute('href')[7:]
                    prof_email = email
                    row[4].value = prof_email
                except:
                    driver.find_element_by_id("edit-last-name").clear()
                    continue
                if len(social) > 1:
                    try:
                        website = driver.find_elements_by_xpath("//*[@id='tau-person-results-table']/tbody/tr/td[5]/div[2]/a")[0].get_attribute('href')
                        prof_website = website
                        row[5].value = prof_website
                    except:
                        driver.find_element_by_id("edit-last-name").clear()
                        continue
                driver.find_element_by_id("edit-last-name").clear()
                wb.save("Results_try_try.xlsx")


                
            
    
def start_scraping():
  driver = InitSelenium()
  driver.get("http://www.ims.tau.ac.il/tal/kr/search_p.aspx")
  courses = scrape_courses(driver, "lstDep6")
  driver.quit()
  return

def write_to_file(worksheet, data, row):
    column = 0
    for item in data:
        worksheet.write(row, column, item)
        column += 1
    row += 1


start_scraping()
    
alfon_adding()